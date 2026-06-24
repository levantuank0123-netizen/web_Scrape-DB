"""Tạo accounts.xlsx với row đầu tiên (Getrewardful - alphana)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "accounts.xlsx"

COLUMNS = [
    ("platform",      "Platform: tolt | getrewardful | affiliatly | firstpromoter | tapfiliate | uppromote | goaffpro | shopify_collabs | refersion | impact"),
    ("label",         "Tên gợi nhớ DUY NHẤT (vd: alphana, brand-abc)"),
    ("dashboard_url", "URL dashboard sau khi login (vd: https://alphana.getrewardful.com/)"),
    ("email",         "Email login (chỉ cần điền nếu profile GPM chưa login dashboard này)"),
    ("password",      "Password login (chỉ cần khi cần auto-login)"),
    ("active",        "TRUE / FALSE"),
    ("notes",         "Ghi chú: brand name, ngày join, ..."),
]

ROWS = [
    ["getrewardful", "alphana", "https://alphana.getrewardful.com/", "tranthu431983@gmail.com", "YomaxGoup1B$", "TRUE", "Test đầu tiên"],
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "accounts"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
for col_idx, (name, _hint) in enumerate(COLUMNS, start=1):
    c = ws.cell(row=1, column=col_idx, value=name)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

hint_font = Font(italic=True, color="666666", size=9)
for col_idx, (_name, hint) in enumerate(COLUMNS, start=1):
    ws.cell(row=2, column=col_idx, value=hint).font = hint_font

for row_idx, row in enumerate(ROWS, start=3):
    for col_idx, val in enumerate(row, start=1):
        ws.cell(row=row_idx, column=col_idx, value=val)

widths = [16, 18, 42, 28, 18, 8, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A3"

wb.save(OUT)
print(f"OK: {OUT}")
