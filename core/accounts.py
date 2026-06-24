"""Đọc accounts.xlsx và trả về list dict tài khoản (GPM mode)."""
from pathlib import Path
import openpyxl

ACCOUNTS_FILE = Path(__file__).parent.parent / "accounts.xlsx"

REQUIRED = ["platform", "label", "dashboard_url"]


def _truthy(v) -> bool:
    if v is None:
        return True
    s = str(v).strip().lower()
    return s in ("", "true", "1", "yes", "y", "có", "co")


def load_accounts(only_active: bool = True, label: str | None = None, platform: str | None = None, owner: str | None = None):
    if not ACCOUNTS_FILE.exists():
        raise FileNotFoundError(
            f"Không tìm thấy {ACCOUNTS_FILE}. Tạo file accounts.xlsx trước."
        )

    wb = openpyxl.load_workbook(ACCOUNTS_FILE, data_only=True)
    ws = wb["accounts"] if "accounts" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    accounts = []
    for row in rows[1:]:
        record = dict(zip(headers, row))
        plat = (record.get("platform") or "").strip().lower() if record.get("platform") else ""
        if not plat:
            continue
        if "|" in plat or " " in plat:
            continue

        record["platform"] = plat
        record["label"] = (record.get("label") or "").strip()
        record["dashboard_url"] = (record.get("dashboard_url") or "").strip()
        record["email"] = (record.get("email") or "").strip() if record.get("email") else ""
        record["password"] = record.get("password") or ""
        record["owner"] = (record.get("owner") or "").strip()
        record["active"] = _truthy(record.get("active"))
        record["notes"] = record.get("notes") or ""

        missing = [c for c in REQUIRED if not record.get(c)]
        if missing:
            print(f"[warn] Bỏ qua row thiếu cột {missing}: label={record.get('label')}")
            continue

        if only_active and not record["active"]:
            continue
        if label and record["label"] != label:
            continue
        if platform and record["platform"] != platform:
            continue
        if owner and record["owner"] != owner:
            continue

        accounts.append(record)

    seen = {}
    for a in accounts:
        key = (a["platform"], a["label"])
        if key in seen:
            raise ValueError(f"Label trùng: platform={a['platform']}, label={a['label']}")
        seen[key] = True

    return accounts
