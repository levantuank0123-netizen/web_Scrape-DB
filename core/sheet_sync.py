"""Đẩy metrics lên Google Sheet qua Apps Script webhook."""
from __future__ import annotations
import json
import urllib.request
import urllib.error
from pathlib import Path

CONFIG_PATH = Path(__file__).parent.parent / "config" / "sheet.json"


def _load_config() -> dict:
    if not CONFIG_PATH.exists():
        raise FileNotFoundError(f"Thiếu {CONFIG_PATH}")
    return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))


def _email_to_password() -> dict:
    """Đọc accounts.xlsx, build mapping email → password."""
    import openpyxl
    accounts_file = Path(__file__).parent.parent / "accounts.xlsx"
    if not accounts_file.exists():
        return {}
    wb = openpyxl.load_workbook(accounts_file, data_only=True)
    ws = wb["accounts"] if "accounts" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    try:
        email_idx = headers.index("email")
        pass_idx = headers.index("password")
    except ValueError:
        return {}
    mapping = {}
    for row in rows[1:]:
        email = (row[email_idx] or "").strip() if row[email_idx] else ""
        password = row[pass_idx] or ""
        if email and password and email not in mapping:
            mapping[email] = password
    return mapping


def _label_to_login_url() -> dict:
    """Đọc accounts.xlsx, build mapping (platform, label) → dashboard_url."""
    import openpyxl
    accounts_file = Path(__file__).parent.parent / "accounts.xlsx"
    if not accounts_file.exists():
        return {}
    wb = openpyxl.load_workbook(accounts_file, data_only=True)
    ws = wb["accounts"] if "accounts" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    try:
        plat_idx = headers.index("platform")
        label_idx = headers.index("label")
        url_idx = headers.index("dashboard_url")
    except ValueError:
        return {}
    mapping = {}
    for row in rows[1:]:
        plat = (row[plat_idx] or "").strip().lower() if row[plat_idx] else ""
        label = (row[label_idx] or "").strip() if row[label_idx] else ""
        url = (row[url_idx] or "").strip() if row[url_idx] else ""
        if plat and label and url:
            mapping[(plat, label)] = url
    return mapping


def push_to_sheet(records: list[dict], timeout: int = 60) -> dict:
    """POST list of records to Apps Script webhook. Trả về {ok, written} hoặc {ok:false, error}."""
    if not records:
        return {"ok": True, "written": 0}

    # Enrich với password + login_url từ accounts.xlsx
    email_to_pass = _email_to_password()
    label_to_url = _label_to_login_url()
    for r in records:
        if not r.get("password"):
            r["password"] = email_to_pass.get(r.get("email", ""), "")
        if not r.get("login_url"):
            r["login_url"] = label_to_url.get((r.get("platform"), r.get("label")), "")

    cfg = _load_config()
    url = cfg["webhook_url"]
    payload = json.dumps(records, ensure_ascii=False).encode("utf-8")

    req = urllib.request.Request(
        url,
        data=payload,
        method="POST",
        headers={"Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8")
            try:
                return json.loads(raw)
            except json.JSONDecodeError:
                return {"ok": False, "error": f"Non-JSON response: {raw[:300]}"}
    except urllib.error.HTTPError as e:
        return {"ok": False, "error": f"HTTP {e.code}: {e.read().decode('utf-8', errors='ignore')[:300]}"}
    except Exception as e:
        return {"ok": False, "error": f"{type(e).__name__}: {e}"}
