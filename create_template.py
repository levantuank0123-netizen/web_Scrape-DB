"""Tạo accounts_template.xlsx (GPM mode — không cần password).

Mỗi dòng = 1 dashboard sẽ được mở trong profile GPM 'affiliate-dashboards'.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "accounts_template_v2.xlsx"

COLUMNS = [
    ("platform",      "tolt | affiliatly | firstpromoter | getrewardful | tapfiliate | uppromote | goaffpro | shopify_collabs | refersion | impact"),
    ("label",         "Tên gợi nhớ duy nhất, vd: 'brand-abc'"),
    ("dashboard_url", "URL trang dashboard (sau khi login). Để trống thì script dùng URL mặc định của platform"),
    ("active",        "TRUE / FALSE — set FALSE để tạm bỏ qua"),
    ("notes",         "Ghi chú: tên brand, ngày join, ..."),
]

SAMPLE = [
    ["tolt", "brand-abc", "https://app.tolt.io/", "TRUE", "Demo Tolt"],
    ["affiliatly", "brand-xyz", "", "FALSE", ""],
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "accounts"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
for col_idx, (name, _hint) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

hint_font = Font(italic=True, color="666666", size=9)
for col_idx, (_name, hint) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=2, column=col_idx, value=hint)
    cell.font = hint_font

for row_idx, row in enumerate(SAMPLE, start=3):
    for col_idx, val in enumerate(row, start=1):
        ws.cell(row=row_idx, column=col_idx, value=val)

widths = [18, 24, 40, 8, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A3"

ws2 = wb.create_sheet("README")
lines = [
    "AFFILIATE DASHBOARD SCRAPER — Template (GPM mode)",
    "",
    "1. Đổi tên file này thành 'accounts.xlsx'.",
    "2. Mỗi dòng = 1 dashboard cần scrape.",
    "3. KHÔNG cần email/password — script connect vào profile GPM 'affiliate-dashboards' đã login sẵn.",
    "4. Cột 'platform': bắt buộc đúng tên (tolt, affiliatly, ...).",
    "5. Cột 'label': DUY NHẤT, dùng để định danh trong DB và Google Sheet.",
    "6. Cột 'dashboard_url': URL chính xác của trang dashboard sau khi login.",
    "7. Cột 'active' = FALSE → bỏ qua row đó.",
    "",
    "QUY TRÌNH:",
    "  - Lần đầu: mở GPM, start profile 'affiliate-dashboards', login thủ công từng dashboard.",
    "  - Sau đó: chạy `python -m core.runner --all` — script tự mở profile, lần lượt vào từng URL, scrape.",
]
for i, line in enumerate(lines, start=1):
    ws2.cell(row=i, column=1, value=line)
ws2.column_dimensions["A"].width = 90

wb.save(OUT)
print(f"OK: {OUT}")
