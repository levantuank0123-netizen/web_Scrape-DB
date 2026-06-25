"""Import dự án từ master Sheet (data/master_import.csv) vào accounts.xlsx.

Hiện tại chỉ import các dự án trên platform đã có adapter (getrewardful).
Có thể chạy lại bất cứ lúc nào — mỗi lần overwrite accounts.xlsx.

Usage:
    python import_master.py                       # import tất cả platform đã hỗ trợ
    python import_master.py --platform getrewardful
    python import_master.py --owner Dũng
"""
from __future__ import annotations
import argparse
import csv
import re
import sys
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MASTER_CSV = Path(__file__).parent / "data" / "master_import.csv"
OUT = Path(__file__).parent / "accounts.xlsx"

# Map domain → platform key trong scrapers registry
DOMAIN_TO_PLATFORM = {
    "getrewardful.com": "getrewardful",
    "firstpromoter.com": "firstpromoter",
    "tolt.io": "tolt",
    "goaffpro.com": "goaffpro",
    # Thêm dần khi có adapter mới:
    # "promotekit.com": "promotekit",
    # "uppromote.com": "uppromote",
    # "trackdesk.com": "trackdesk",
    # "partnerstack.com": "partnerstack",
    # "dub.co": "dub",
    # "impact.com": "impact",
}

COLUMNS = [
    "platform", "label", "dashboard_url", "email", "password",
    "owner", "active", "notes",
]


def detect_platform(login_url: str) -> str | None:
    host = urlparse(login_url).netloc.lower()
    for domain, key in DOMAIN_TO_PLATFORM.items():
        if domain in host:
            return key
    return None


def slugify(name: str) -> str:
    s = name.strip().lower()
    s = re.sub(r"[^\w]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s or "noname"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--platform", help="Chỉ import 1 platform key (vd: getrewardful)")
    ap.add_argument("--owner", help="Chỉ import 1 người quản (Dũng/An/Hà)")
    ap.add_argument("--limit", type=int, help="Giới hạn số rows để test")
    args = ap.parse_args()

    if not MASTER_CSV.exists():
        print(f"Không tìm thấy {MASTER_CSV}")
        return 1

    imported = []
    skipped_no_adapter = []

    with open(MASTER_CSV, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            login = (r.get("Affiliate Login") or "").strip()
            project = (r.get("Project") or "").strip()
            if not login or not project:
                continue

            plat = detect_platform(login)
            if not plat:
                skipped_no_adapter.append((project, login))
                continue

            if args.platform and plat != args.platform:
                continue

            owner = (r.get("Account Reg") or "").strip()
            if args.owner and owner != args.owner:
                continue

            imported.append({
                "platform": plat,
                "label": slugify(project),
                "dashboard_url": login,
                "email": (r.get("Email") or "").strip(),
                "password": r.get("Pass Affiliate") or "",
                "owner": owner,
                "active": "TRUE",
                "notes": f"Type={r.get('Type','')} | ref={(r.get('Link Ref') or '').strip()}",
            })

    # Loại label trùng (giữ row đầu)
    seen = set()
    unique = []
    dupes = []
    for r in imported:
        key = (r["platform"], r["label"])
        if key in seen:
            dupes.append(r["label"])
            continue
        seen.add(key)
        unique.append(r)

    if args.limit:
        unique = unique[:args.limit]

    # Write to xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "accounts"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for col_idx, name in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for row_idx, rec in enumerate(unique, start=2):
        for col_idx, name in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=rec.get(name, ""))

    widths = [16, 28, 42, 30, 18, 10, 8, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(OUT)

    import sys
    sys.stderr.write(f"\nOK: ghi {len(unique)} dong vao {OUT}\n")
    sys.stderr.write(f"  - Skipped (chua co adapter): {len(skipped_no_adapter)} du an\n")
    if dupes:
        sys.stderr.write(f"  - Bo qua duplicate label: {len(dupes)}\n")

    from collections import Counter
    plat_count = Counter(r["platform"] for r in unique)
    owner_count = Counter(r["owner"] for r in unique)
    sys.stderr.write(f"\n--- Phan bo ---\n")
    sys.stderr.write(f"Platform: {dict(plat_count)}\n")
    sys.stderr.write(f"Owner: {dict(owner_count)}\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
